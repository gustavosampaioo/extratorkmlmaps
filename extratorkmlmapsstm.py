import os
import streamlit as st
from zipfile import ZipFile
import xml.etree.ElementTree as ET
import openpyxl
import csv
from collections import defaultdict
import webbrowser

class KMLProcessorApp:
    def __init__(self, root):
        self.root = root
        self.arquivo_kml = None
        self.arquivo_db = None
        self.arquivo_db_cto = None




    def carregar_kml(self):
        arquivo_kml = st.file_uploader(
            title='Escolha o arquivo KML ou KMZ',
            filetypes=[('KML Files', '*.kml'), ('KMZ Files', '*.kmz')]
        )

        if arquivo_kml:
            self.arquivo_kml = arquivo_kml
            st.write(f"Arquivo KML carregado: {arquivo_kml.name}")

    def carregar_db(self):
        arquivo_db = st.file_uploader(
            "Escolha o arquivo do banco de dados Clientes (Excel)",
            filetypes=[('Excel Files', '*.xlsx')]
        )

        if arquivo_db:
            self.arquivo_db = arquivo_db
            wb = openpyxl.load_workbook(arquivo_db)
            sheet = wb.active
            dados_db = defaultdict(list)
            for excel_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                nome_tratado = excel_row[0].split(' ')[0]
                ids = excel_row[6]
                status = excel_row[4]  # Supondo que o status está na coluna 5
                dados_db[nome_tratado].append({
                    'id': ids,
                    'status': status  # Armazenando o status
                })

            st.write(f"Arquivo DB Clientes carregado: {arquivo_db.name}")


    def carregar_db_cto(self):
        arquivo_db_cto = st.file_uploader(
            "Escolha o arquivo do banco de dados CTO (Excel)",
            type=["xlsx"]
        )
        if arquivo_db_cto:
            self.arquivo_db_cto = arquivo_db_cto
            st.write(f"Arquivo DB CTO carregado: {arquivo_db_cto.name}")

    def processar(self):
        if self.arquivo_kml and self.arquivo_db and self.arquivo_db_cto:
            if self.arquivo_kml.name.lower().endswith('.kmz'):
                kml_files = self.extrair_marcadores()
                if not kml_files:
                    st.warning("Nenhum arquivo KML encontrado no KMZ.")
                else:
                    for kml_file in kml_files:
                        with ZipFile(self.arquivo_kml, 'r') as zip_ref:
                            kml_content = zip_ref.read(kml_file).decode('utf-8')
                            dados_kml = self.extrair_dados_kml(kml_content)
                            dados_cto = self.extrair_dados_cto()
                            self.atualizar_descricao_com_db(dados_kml, dados_cto)
                            self.gerar_kml_csv_atualizado(kml_file, dados_kml)
                            link = self.gerar_link_google_maps(dados_kml[0][1], dados_kml[0][2])
                            st.markdown(f"[Abrir no Google Maps]({link})", unsafe_allow_html=True)
            elif self.arquivo_kml.name.lower().endswith('.kml'):
                kml_content = self.arquivo_kml.read().decode('utf-8')
                dados_kml = self.extrair_dados_kml(kml_content)
                dados_cto = self.extrair_dados_cto()
                self.atualizar_descricao_com_db(dados_kml, dados_cto)
                self.gerar_kml_csv_atualizado(self.arquivo_kml.name, dados_kml)
            else:
                st.warning("Formato de arquivo não suportado. Por favor, selecione um arquivo KML ou KMZ.")
            st.success("Extração concluída")


    def extrair_dados_cto(self):
        wb_cto = openpyxl.load_workbook(self.arquivo_db_cto)
        sheet_cto = wb_cto.active
        dados_cto = {}

        nome_pon_set = set()

        for excel_row in sheet_cto.iter_rows(min_row=2, max_row=sheet_cto.max_row, values_only=True):
            nome_cto_pon_parts = excel_row[0].split(' ', 1)
            nome_sigla = excel_row[0].split('-', 1)[0].strip()
            if len(nome_cto_pon_parts) == 2:
                nome_cto, nome_pon = nome_cto_pon_parts
                nome_pon_set.add(nome_pon)
            else:
                nome_cto = excel_row[0]
                nome_pon = ""

            description_cto = excel_row[3]
            total_portas_utilizadas = excel_row[4] if excel_row[4] is not None else 0  # Verificação e substituição por 0
            total_portas_disponiveis = excel_row[5] if excel_row[5] is not None else 0  # Verificação e substituição por 0
            dados_cto[nome_cto] = {
                'description_splitter': description_cto,
                'total_portas_utilizadas': total_portas_utilizadas,
                'total_portas_disponiveis': total_portas_disponiveis,
                'nome_cto': nome_cto,
                'nome_pon': nome_pon,
                'nome_sigla': nome_sigla
            }

        soma_coluna_5 = 0

        for nome_pon in nome_pon_set:
            soma_coluna_5 += sum(int(info['total_portas_utilizadas']) for info in dados_cto.values() if
                                info.get('nome_pon') == nome_pon)

        soma_portas_por_nome_pon = self.calcular_soma_portas_por_nome_pon(dados_cto, self.arquivo_kml)

        dados_cto['soma_portas_por_nome_pon'] = soma_portas_por_nome_pon
        dados_cto['soma_coluna_5'] = soma_coluna_5

        return dados_cto


    def calcular_soma_servicos_por_nome_pon(self, dados_cto, arquivo_db, arquivo_kml):
        # Função para extrair o nome do placemark do KML
        def extrair_nome_placemark(arquivo_kml):
            try:
                tree = ET.parse(arquivo_kml)
                root = tree.getroot()

                nomes_placemarks = []

                for placemark in root.findall('.//{http://www.opengis.net/kml/2.2}Placemark'):
                    nome_element = placemark.find('.//{http://www.opengis.net/kml/2.2}name')

                    if nome_element is not None:
                        nome_placemark = nome_element.text.split('-')[0].strip()
                        nomes_placemarks.append(nome_placemark)

                return nomes_placemarks

            except Exception as e:
                print(f"Erro ao extrair nomes dos placemarks: {e}")
                return []

        # Extrair o nome do placemark do KML
        nomes_placemarks_kml = extrair_nome_placemark(arquivo_kml)

        # Carregar o arquivo de banco de dados
        wb_db = openpyxl.load_workbook(arquivo_db)
        sheet_db = wb_db.active
        dados_servicos = {}

        # Iterar sobre as linhas do arquivo de banco de dados
        for excel_row in sheet_db.iter_rows(min_row=2, max_row=sheet_db.max_row, values_only=True):
            # Extrair o nome do cliente e o tipo de serviço da linha do banco de dados
            nome_cliente = excel_row[0].split(' ')[0]
            descricao_servico = excel_row[3]  # Coluna 4 contém a descrição do serviço

            # Extrair o tipo de serviço da descrição
            tipo_servico = descricao_servico

            # Extrair o nome do PON a partir do nome do cliente
            nome_prefixo = nome_cliente.split('-')[0].strip()

            # Verificar se o nome do prefixo está presente nos nomes dos placemarks do KML
            if nome_prefixo in nomes_placemarks_kml:
                # Obter o nome do PON e a sigla associados ao cliente
                nome_pon = dados_cto.get(nome_cliente, {}).get('nome_pon', '')
                nome_sigla = dados_cto.get(nome_cliente, {}).get('nome_sigla', '')

                # Usar a sigla para o prefixo
                if nome_pon and nome_sigla:
                    chave = (nome_pon, nome_sigla)  # Usar uma tupla como chave

                    # Inicializar o contador de serviços para a chave se ainda não existir
                    if chave not in dados_servicos:
                        dados_servicos[chave] = defaultdict(int)

                    # Incrementar a quantidade para o tipo de serviço correspondente
                    dados_servicos[chave][tipo_servico] += 1

        # Imprimir no terminal os valores de "nome_pon" e "nome_sigla"
        for (nome_pon, nome_sigla), servicos in dados_servicos.items():
            print(f"Nome_PON: {nome_pon}, Nome_Sigla: {nome_sigla}, Serviços: {dict(servicos)}")

        return dados_servicos


    # Função para calcular a soma das portas por nome_pon
    def calcular_soma_portas_por_nome_pon(self, dados_cto, arquivo_kml):
        # Função para extrair o nome do placemark do KML
        def extrair_nome_placemark(arquivo_kml):
            try:
                tree = ET.parse(arquivo_kml)
                root = tree.getroot()

                nomes_placemarks = []

                for placemark in root.findall('.//{http://www.opengis.net/kml/2.2}Placemark'):
                    nome_element = placemark.find('.//{http://www.opengis.net/kml/2.2}name')

                    if nome_element is not None:
                        nome_placemark = nome_element.text.split('-')[0].strip()
                        nomes_placemarks.append(nome_placemark)

                return nomes_placemarks

            except Exception as e:
                print(f"Erro ao extrair nomes dos placemarks: {e}")
                return []

        # Extrair o nome do placemark do KML
        nomes_placemarks_kml = extrair_nome_placemark(arquivo_kml)

        soma_portas_por_nome_pon = {}

        # Definindo os status que queremos considerar
        status_permitidos = {
            "Serviço Habilitado",
            "Suspenso a Pedido do Cliente",
            "Suspenso Parcialmente",
            "Suspenso por Débito"
        }

        for info in dados_cto.values():
            if isinstance(info, dict):  # Verifique se info é um dicionário
                nome_pon = info.get('nome_pon', '')
                nome_sigla = info.get('nome_sigla', '')  # Obter a sigla da CTO

                if nome_pon and nome_sigla:
                    nome_prefixo = nome_sigla.split('-', 1)[0].strip()  # Usar a sigla para o prefixo

                    if nome_prefixo in nomes_placemarks_kml:
                        chave = (nome_pon, nome_sigla)  # Usar uma tupla como chave
                        if chave not in soma_portas_por_nome_pon:
                            soma_portas_por_nome_pon[chave] = {
                                'total_portas': 0,
                                'nome_tratado_list': []
                            }

                        # Verificar o status na base de clientes
                        nome_tratado = info.get('nome_cto', '').split(' ')[0]  # Ajuste conforme necessário
                        clientes = self.dados_db.get(nome_tratado, [])
                        
                        # Debugging: Verifique se estamos acessando os clientes corretamente
                        print(f"Verificando clientes para {nome_tratado}: {clientes}")

                        # Flag para verificar se algum cliente tem status permitido
                        status_permitido_encontrado = False

                        for cliente in clientes:
                            print(f"Verificando cliente: {cliente}")  # Debugging
                            if cliente['status'] in status_permitidos:
                                status_permitido_encontrado = True
                                break  # Se já encontramos um status permitido, não precisamos continuar verificando

                        # Se algum cliente tem status permitido, somamos as portas
                        if status_permitido_encontrado:
                            total_portas_utilizadas = int(info.get('total_portas_utilizadas', 0))
                            soma_portas_por_nome_pon[chave]['total_portas'] += total_portas_utilizadas
                            soma_portas_por_nome_pon[chave]['nome_tratado_list'].append(info.get('nome_cto', ''))  # Se precisar do nome completo da CTO

        # Imprimir no terminal os valores de "nome_pon" e "nome_sigla"
        for (nome_pon, nome_sigla), info in soma_portas_por_nome_pon.items():
            print(f"Nome_PON: {nome_pon}, Nome_Sigla: {nome_sigla}, Total_Portas: {info['total_portas']}")
            print(f"Nome_CTOs: {', '.join(info['nome_tratado_list'])}")

        # Aqui, total_clientes_pon já está considerando apenas os status permitidos
        total_clientes_pon = {}
        for (nome_pon, nome_sigla), info in soma_portas_por_nome_pon.items():
            if info['total_portas'] > 0:  # Somente incluir se houver portas somadas
                total_clientes_pon[(nome_pon, nome_sigla)] = info['total_portas']

        return total_clientes_pon


    def extrair_marcadores(self):
        kml_files = []
        with ZipFile(self.arquivo_kml, 'r') as zip_ref:
            for file_info in zip_ref.infolist():
                if file_info.filename.lower().endswith('.kml'):
                    kml_files.append(file_info.filename)
        return kml_files

    def extrair_dados_kml(self, kml_content):
        root = ET.fromstring(kml_content)
        dados = []
        for placemark in root.findall(".//{http://www.opengis.net/kml/2.2}Placemark"):
            nome_element = placemark.find(".//{http://www.opengis.net/kml/2.2}name")
            coordenadas_element = placemark.find(".//{http://www.opengis.net/kml/2.2}coordinates")
            description_element = placemark.find(".//{http://www.opengis.net/kml/2.2}description")

            if nome_element is not None and coordenadas_element is not None and description_element is not None:
                nome = nome_element.text
                coordenadas = coordenadas_element.text
                description = description_element.text

                try:
                    longitude, latitude, _ = map(str, coordenadas.split(','))
                    nome_parte = nome.split(' ')[0]
                    nome_sigla = nome.split('-', 1)[0].strip()  # Adicionado para extrair a sigla
                    nome_pon = self.obter_nome_pon_a_partir_do_nome_cto(nome_parte)

                    dados.append([nome_parte, latitude, longitude, description, nome_pon, nome_sigla])  # Incluído nome_sigla nos dados
                except ValueError as e:
                    print(f"Erro ao converter coordenadas para '{nome}': {coordenadas}. Erro: {e}")
            else:
                print("Elemento ausente em um placemark. Ignorando.")

        return dados


    def obter_nome_pon_a_partir_do_nome_cto(self, nome_cto):
        partes_nome_cto = nome_cto.split(' ')
        if len(partes_nome_cto) > 1:
            nome_pon = partes_nome_cto[1]
            return nome_pon
        else:
            return None

    def atualizar_descricao_com_db(self, dados_kml, dados_cto):
        wb = openpyxl.load_workbook(self.arquivo_db)
        sheet = wb.active
        dados_db = defaultdict(list)

        # Obtendo apenas as informações necessárias do arquivo de banco de dados
        for excel_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            dados_db[excel_row[0].split(' ')[0]].append({
                'id': excel_row[6],
                'nome_pon': excel_row[1],  # Assumindo que a coluna 1 contém o nome PON
                'servico': excel_row[3],  # Serviço associado, extraído da coluna 4
                'status': excel_row[4]  # Status associado, extraído da coluna 5
            })

        # Calcular a soma dos serviços por PON
        soma_servicos = self.calcular_soma_servicos_por_nome_pon(dados_cto, self.arquivo_db, self.arquivo_kml)

        # Obter a soma total de clientes por PON
        total_clientes_pon = self.calcular_soma_portas_por_nome_pon(dados_cto, self.arquivo_kml)

        for row in dados_kml:
            nome = row[0]
            nome_tratado = nome.split(' ')[0]
            nova_descricao = ""
            descricao_cto_info = dados_cto.get(nome_tratado, {})

            if 'description_splitter' in descricao_cto_info:
                nova_descricao += f"Splitter de Atendimento 1x{descricao_cto_info['description_splitter']}\n\n"
                nova_descricao += f"Portas da CTO Utilizadas: {descricao_cto_info.get('total_portas_utilizadas', 'N/A')}\n"
                nova_descricao += f"Portas da CTO Disponíveis: {descricao_cto_info.get('total_portas_disponiveis', 'N/A')}\n\n"

                # Filtrando apenas as informações relevantes do banco de dados
                info_db = dados_db.get(nome_tratado, [])

                if info_db:
                    nova_descricao += "Listagem de IDs com Serviço:\n"
                    for entry in info_db:
                        id_formatado = entry['id'].ljust(9)  # Ajustar o ID para ocupar 9 caracteres, adicionando espaços à direita, se necessário
                        nova_descricao += f" - ID: {id_formatado}, Status: {entry['status']}\n"
                else:
                    nova_descricao += "Nenhum Id Ativo\n"

                # Obter a soma de clientes na PON usando a sigla
                nome_sigla = descricao_cto_info.get('nome_sigla', '')
                soma_portas = total_clientes_pon.get((descricao_cto_info['nome_pon'], nome_sigla), 'N/A')
                nova_descricao += f"\nTotal de Clientes na PON: {soma_portas}\n"  # Atualizado para usar total_portas

                # Incluir a soma dos serviços por PON na descrição
                nova_descricao += "\nServiços na PON:\n"
                if (descricao_cto_info['nome_pon'], nome_sigla) in soma_servicos:
                    for tipo_servico, quantidade in soma_servicos[(descricao_cto_info['nome_pon'], nome_sigla)].items():
                        nova_descricao += f" - {tipo_servico}: {quantidade}\n"

            else:
                nova_descricao += "Informações CTO não disponíveis.\n"

            row[3] = self.formatar_descricao_vertical(nova_descricao)

        wb.save(self.arquivo_db)



    def substituir_description_kml(self, kml_file, dados_kml):
        tree = ET.parse(kml_file)
        root = tree.getroot()

        namespace = {'kml': 'http://www.opengis.net/kml/2.2'}

        # Encontrar todas as pastas "BACKBONE"
        backbone_folders = []
        for folder in root.findall(".//kml:Folder", namespace):
            name_element = folder.find("kml:name", namespace)
            if name_element is not None and name_element.text == "BACKBONE":
                backbone_folders.append(folder)

        # Criar um conjunto de placemarks dentro de todas as pastas "BACKBONE"
        placemarks_backbone = set()
        for backbone_folder in backbone_folders:
            for placemark in backbone_folder.findall(".//kml:Placemark", namespace):
                name_element = placemark.find("kml:name", namespace)
                if name_element is not None:
                    placemarks_backbone.add(name_element.text)

        # Agora percorre todos os placemarks no KML e substitui as descrições, se necessário
        for placemark in root.findall(".//kml:Placemark", namespace):
            nome_placemark = placemark.find("kml:name", namespace)
            if nome_placemark is not None:
                nome_placemark_text = nome_placemark.text.split(' ')[0]

                # Verificar se o placemark está dentro de alguma pasta "BACKBONE"
                if nome_placemark.text in placemarks_backbone:
                    continue  # Pula a atualização da descrição para este placemark

                # Substituir as descrições se corresponder ao nome
                for dados_row in dados_kml:
                    if nome_placemark_text == dados_row[0]:
                        description = placemark.find("kml:description", namespace)
                        if description is not None:
                            # Substituir o conteúdo da descrição
                            description.text = self.formatar_descricao_vertical(dados_row[3])
                            description.text += "\n"
                            latitude, longitude = dados_row[1], dados_row[2]
                            description.text += f"Localização: {self.gerar_link_google_maps(latitude, longitude)}"

                        # Atualizar descrição em Point e LineString, se existirem
                        point = placemark.find("kml:Point", namespace)
                        if point is not None:
                            point_description = point.find("kml:description", namespace)
                            if point_description is not None:
                                point_description.text = self.formatar_descricao_vertical(dados_row[3])

                        linestring = placemark.find("kml:LineString", namespace)
                        if linestring is not None:
                            linestring_description = linestring.find("kml:description", namespace)
                            if linestring_description is not None:
                                linestring_description.text = self.formatar_descricao_vertical(dados_row[3])

        return tree



    def formatar_descricao_vertical(self, descricao):
        return "<pre>" + descricao.rstrip('\n') + "</pre>"

    def gerar_kml_csv_atualizado(self, kml_file, dados):
        tree = self.substituir_description_kml(kml_file, dados)
        novo_kml_file = f'{os.path.splitext(kml_file)[0]}_atualizado.kml'
        tree.write(novo_kml_file, encoding='utf-8', xml_declaration=True)

        csv_file = f'{os.path.splitext(kml_file)[0]}_atualizado.csv'
        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerow(['Nome', 'Latitude', 'Longitude', 'CTO', 'IDs'])
            for row in dados:
                nome_tratado = row[0]
                latitude = row[1]
                longitude = row[2]
                cto_ids = row[3].replace('\n', '_')
                csv_writer.writerow([nome_tratado, latitude, longitude, cto_ids])

        print(f'Dados extraídos e descrição atualizada para {csv_file}')
        print(f'Arquivo KML atualizado salvo como {novo_kml_file}')

        with open(csv_file, 'r', encoding='utf-8') as csv_read:
            print("\nConteúdo do arquivo CSV:")
            print(csv_read.read())

    def gerar_link_google_maps(self, latitude, longitude):
        google_maps_url = f'https://www.google.com/maps/place/{latitude},{longitude}'
        geo_uri = f'<a href="{google_maps_url}" target="_blank">Abrir no Google Maps</a>'
        return geo_uri

def main():
    st.title("ATUALIZADOR DE MAPAS KML")
    app = KMLProcessorApp()
    app.carregar_kml()
    app.carregar_db()
    app.carregar_db_cto()
    if st.button("Processar"):
        app.processar()

if __name__ == "__main__":
    main()
